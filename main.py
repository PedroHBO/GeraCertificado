from datetime import datetime
import locale
from docx import Document
from openpyxl import load_workbook
import docx.shared
import shutil

# Carregar o arquivo Excel
workbook = load_workbook(
    'C:/Users/pedro.oliveira/Desktop/CertificadoTCS/TCS.xlsx')
sheet = workbook['TCS']

# Carregar o arquivo Word template
caminho_word = 'C:/Users/pedro.oliveira/Desktop/CertificadoTCS/templateTCS.docx'
endereco = 'C:/Users/pedro.oliveira/Desktop/CertificadoTCS/TCS/'

# Obter a data atual
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
data_atual = datetime.now().strftime('%d de %B de %Y')

# Copiar o template para cada usuário
for row in range(2, sheet.max_row + 1):
    projeto = sheet.cell(row=row, column=1).value
    autores = sheet.cell(row=row, column=2).value
    mentoria = sheet.cell(row=row, column=3).value

    # Verificar se o nome não é nulo
    if autores is not None:
        # Copiar o template para um novo arquivo
        autor = autores.replace(" ", "_").replace(".", "").replace("\n", "")
        caminho_word_custumizado = f'{endereco}Certificado_{autor}.docx'
        shutil.copy(caminho_word, caminho_word_custumizado)

        # Abrir o documento copiado
        doc = Document(caminho_word_custumizado)

        # Preencher o documento com os dados do Excel
        for paragraph in doc.paragraphs:
            paragraph.text = paragraph.text.replace('[projeto]', projeto)\
                .replace('[nome]', autores)\
                .replace('[mentor]', mentoria)\
                .replace('[data]', data_atual)
            for run in paragraph.runs:
                font = run.font
                font.name = 'Arial'
                font.size = docx.shared.Pt(20)
                font.color.rgb = docx.shared.RGBColor(122, 23, 18)

        # Salvar o documento com os dados preenchidos
        doc.save(caminho_word_custumizado)
